"""
Excel Extractor — convert .xlsx / .xls / .csv to markdown table.

Multi-sheet files produce one ## Sheet: {name} section per sheet.
Merge cells are forward-filled so no values are lost.
Empty rows and columns are stripped before rendering.
"""

import csv
import io
import re
from pathlib import Path


class ExcelExtractor:
    MAX_ROWS = 100_000  # per-sheet row limit
    MAX_CELLS = 500_000  # per-sheet cell limit (rows * columns) to bound memory

    def extract(self, file_bytes: bytes, file_name: str) -> tuple[str, dict]:
        suffix = Path(file_name).suffix.lower()
        if suffix in (".xlsx", ".xls"):
            md, meta = self._extract_excel(file_bytes, suffix)
        elif suffix == ".csv":
            md, meta = self._extract_csv(file_bytes, file_name)
        else:
            raise ValueError(f"Unsupported Excel format: {suffix}")

        if not md.strip():
            raise ValueError("No data extracted from Excel file")

        return md, meta

    # ------------------------------------------------------------------
    # .xlsx / .xls
    # ------------------------------------------------------------------

    def _extract_excel(self, file_bytes: bytes, suffix: str) -> tuple[str, dict]:
        if suffix == ".xlsx":
            return self._extract_xlsx(file_bytes)
        return self._extract_xls(file_bytes)

    def _extract_xlsx(self, file_bytes: bytes) -> tuple[str, dict]:
        import openpyxl

        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
        sections: list[str] = []
        sheet_names: list[str] = []

        try:
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                if ws.max_row and ws.max_row > self.MAX_ROWS:
                    raise ValueError(
                        f"Sheet '{sheet_name}' has {ws.max_row} rows, "
                        f"which exceeds the limit of {self.MAX_ROWS}"
                    )
                est_cells = (ws.max_row or 0) * (ws.max_column or 0)
                if est_cells > self.MAX_CELLS:
                    raise ValueError(
                        f"Sheet '{sheet_name}' has ~{est_cells} cells, "
                        f"which exceeds the limit of {self.MAX_CELLS}"
                    )
                rows = self._read_sheet_rows(ws)
                rows = self._strip_empty(rows)
                if not rows:
                    continue
                sheet_names.append(sheet_name)
                sections.append(self._render_sheet(sheet_name, rows))
        finally:
            wb.close()

        if not sections:
            return "", {"sheet_count": 0, "sheets": []}

        meta = {"sheet_count": len(sections), "sheets": sheet_names}
        return "\n\n".join(sections), meta

    def _read_sheet_rows(self, ws) -> list[list[str]]:
        """Read all rows, forward-filling merge cells."""
        merged_map = self._build_merged_map(ws)
        rows: list[list[str]] = []

        for row_idx, row in enumerate(ws.iter_rows(), start=1):
            cells: list[str] = []
            for col_idx, cell in enumerate(row, start=1):
                value = merged_map.get((row_idx, col_idx))
                if value is None:
                    value = cell.value
                cells.append(self._cell_to_str(value))
            rows.append(cells)

        return rows

    def _build_merged_map(self, ws) -> dict[tuple[int, int], str]:
        """Build mapping from every cell in a merged range → top-left value."""
        merged_map: dict[tuple[int, int], str] = {}

        for merged_range in ws.merged_cells.ranges:
            top_left_value = ws.cell(
                row=merged_range.min_row, column=merged_range.min_col
            ).value
            fill_value = self._cell_to_str(top_left_value)
            for row in range(merged_range.min_row, merged_range.max_row + 1):
                for col in range(merged_range.min_col, merged_range.max_col + 1):
                    merged_map[(row, col)] = fill_value

        return merged_map

    def _extract_xls(self, file_bytes: bytes) -> tuple[str, dict]:
        import xlrd

        wb = xlrd.open_workbook(file_contents=file_bytes)
        sections: list[str] = []
        sheet_names: list[str] = []

        try:
            for sheet in wb.sheets():
                if sheet.nrows > self.MAX_ROWS:
                    raise ValueError(
                        f"Sheet '{sheet.name}' has {sheet.nrows} rows, "
                        f"which exceeds the limit of {self.MAX_ROWS}"
                    )
                if sheet.nrows * sheet.ncols > self.MAX_CELLS:
                    raise ValueError(
                        f"Sheet '{sheet.name}' has {sheet.nrows * sheet.ncols} cells, "
                        f"which exceeds the limit of {self.MAX_CELLS}"
                    )
                merged_map = self._build_xls_merged_map(sheet)
                rows: list[list[str]] = []
                for row_idx in range(sheet.nrows):
                    cells: list[str] = []
                    for col_idx in range(sheet.ncols):
                        value = merged_map.get((row_idx, col_idx))
                        if value is None:
                            value = sheet.cell_value(row_idx, col_idx)
                        cells.append(self._cell_to_str(value))
                    rows.append(cells)
                rows = self._strip_empty(rows)
                if not rows:
                    continue
                sheet_names.append(sheet.name)
                sections.append(self._render_sheet(sheet.name, rows))
        finally:
            wb.release_resources()

        if not sections:
            return "", {"sheet_count": 0, "sheets": []}

        meta = {"sheet_count": len(sections), "sheets": sheet_names}
        return "\n\n".join(sections), meta

    @staticmethod
    def _build_xls_merged_map(sheet) -> dict[tuple[int, int], str]:
        """Build mapping from every cell in merged ranges to the top-left value (xlrd)."""
        merged_map: dict[tuple[int, int], str] = {}
        for rlo, rhi, clo, chi in sheet.merged_cells:
            top_left = sheet.cell_value(rlo, clo)
            fill = ExcelExtractor._cell_to_str(top_left)
            for r in range(rlo, rhi):
                for c in range(clo, chi):
                    merged_map[(r, c)] = fill
        return merged_map

    # ------------------------------------------------------------------
    # CSV
    # ------------------------------------------------------------------

    def _extract_csv(self, file_bytes: bytes, file_name: str) -> tuple[str, dict]:
        text = self._detect_and_decode(file_bytes)
        try:
            sample = text[:65536] if len(text) > 1 else "a,b"
            dialect = csv.Sniffer().sniff(sample)
        except csv.Error:
            dialect = "excel"
        reader = csv.reader(io.StringIO(text), dialect)
        rows = [row for row in reader]
        rows = self._strip_empty(rows)
        if not rows:
            return "", {"sheet_count": 0, "sheets": []}
        sheet_name = Path(file_name).stem
        md = self._render_sheet(sheet_name, rows)
        return md, {"sheet_count": 1, "sheets": [sheet_name]}

    def _detect_and_decode(self, file_bytes: bytes) -> str:
        # BOM
        if file_bytes[:3] == b"\xef\xbb\xbf":
            return file_bytes[3:].decode("utf-8")
        if file_bytes[:2] == b"\xff\xfe":
            return file_bytes[2:].decode("utf-16-le")
        if file_bytes[:2] == b"\xfe\xff":
            return file_bytes[2:].decode("utf-16-be")

        # Try UTF-8
        try:
            return file_bytes.decode("utf-8")
        except UnicodeDecodeError:
            pass

        # Try Windows-1258 (Vietnamese)
        try:
            return file_bytes.decode("windows-1258")
        except UnicodeDecodeError:
            pass

        # Fallback latin-1 (never fails)
        return file_bytes.decode("latin-1")

    # ------------------------------------------------------------------
    # Shared helpers
    # ------------------------------------------------------------------

    @staticmethod
    def _cell_to_str(value) -> str:
        if value is None:
            return ""
        if isinstance(value, float):
            if value == int(value):
                return str(int(value))
            return str(value)
        text = str(value).replace("\r\n", " ").replace("\n", " ").replace("|", "\\|")
        return text.strip()

    @staticmethod
    def _strip_empty(rows: list[list[str]]) -> list[list[str]]:
        """Remove fully-empty rows and columns."""
        if not rows:
            return rows

        # Strip empty rows
        rows = [r for r in rows if any(c.strip() for c in r)]
        if not rows:
            return []

        # Strip empty columns (all rows have empty string at that column index)
        max_cols = max(len(r) for r in rows)
        non_empty_cols: list[int] = []
        for col_idx in range(max_cols):
            if any(
                col_idx < len(r) and r[col_idx].strip() for r in rows
            ):
                non_empty_cols.append(col_idx)

        if not non_empty_cols:
            return []

        return [
            [r[col_idx] if col_idx < len(r) else "" for col_idx in non_empty_cols]
            for r in rows
        ]

    @staticmethod
    def _render_sheet(sheet_name: str, rows: list[list[str]]) -> str:
        lines: list[str] = [f"## Sheet: {sheet_name}", ""]

        if not rows:
            return "\n".join(lines)

        max_cols = max(len(r) for r in rows)

        # Normalize all rows to same width
        padded = [r + [""] * (max_cols - len(r)) for r in rows]

        # Header + separator
        header = padded[0]
        lines.append("| " + " | ".join(header) + " |")
        lines.append("|" + "|".join(["-------"] * max_cols) + "|")

        # Data rows
        for row in padded[1:]:
            lines.append("| " + " | ".join(row) + " |")

        return "\n".join(lines)
